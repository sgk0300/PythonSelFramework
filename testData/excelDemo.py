import openpyxl
Dict={}
book = openpyxl.load_workbook("C:\\Users\\SANJOG\\Desktop\\PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1,column=2)
print(cell.value)
'''sheet.cell(row=2,column=3).value="Kulkarni"
print(sheet.cell(row=2,column=3).value)
book.save("PythonDemo.xlsx")'''

print(sheet.max_row)
print(sheet.max_column)
print(sheet['B2'].value)

print(("########################"))
for i in range(1,sheet.max_row+1):
    #if sheet.cell(row=i,column=1).value=="TestCase1":
    for j in range(2,sheet.max_column+1):
        Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i, column=j).value
        '''FirstName = sheet.cell(row=i, column=1).value
        LastName = sheet.cell(row=i, column=2).value
        Gender = sheet.cell(row=i, column=3).value'''

print(Dict)